import os
from email.message import EmailMessage
import ssl
import smtplib
import openpyxl
import imghdr
from PIL import Image
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os
from dotenv import load_dotenv

load_dotenv()

wb = openpyxl.load_workbook('List.xlsx')
sh = wb.active

email_sender = 'neoz.blockchain@gmail.com'
# email_receiver = 'neoz.blockchain@gmail.com'
# email_password = os.environ.get("CR_mail_password_py")
email_password = os.getenv('mail_passcode')

subj = 'Wanna grab some icecream?'
body = '''
Let's get some icecream after college
'''

em = MIMEMultipart()
em['From'] = email_sender
em['Subject'] = subj
# em.attach(body)
em.attach(MIMEText(body))

# with open('fas quote.jpg', 'rb') as f:
#     file_data = f.read()
#     file_type = imghdr.what(f.name)
#     file_name = f.name

# em.add_attachment(file_data, maintype='image', subtype=file_type, filename=file_name)

context = ssl.create_default_context()

for i in range(2,5):

    email_receiver = sh.cell(row=i,column=2).value
    email_receiver_name = sh.cell(row=i,column=1).value

    img_path = f'imagess/{email_receiver_name}.jpg'
    img = Image.open(img_path)
    img_fin = img.convert('RGB')
    pdf_path = f'pic_to_pdf/{email_receiver_name}.pdf'
    img_fin.save(pdf_path)

    binary_pdf = open(pdf_path, 'rb')
    payload = MIMEBase('application','octane-stream',Name=pdf_path)
    payload.set_payload((binary_pdf).read())
    encoders.encode_base64(payload)
    payload.add_header('Content-Decomposition','attachment',filename=pdf_path)
    em.attach(payload)
    binary_pdf.close()
    em['To'] = email_receiver
    # with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
    with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
        smtp.starttls()
        smtp.login(email_sender, email_password)
        smtp.sendmail(email_sender, email_receiver, em.as_string())
        print(f'Message sent to {email_receiver_name}')
        os.remove(pdf_path)

    del em['To']