from email.message import EmailMessage
import ssl
import smtplib
from PIL import Image
import os
import openpyxl
from email.mime.multipart import MIMEMultipart
import os
from dotenv import load_dotenv

load_dotenv()

wb = openpyxl.load_workbook('List.xlsx')
sh = wb.active

email_sender = 'neoz.blockchain@gmail.com'
email_password = os.getenv('mail_passcode')

em = EmailMessage()
# em = MIMEMultipart()

subj = 'Wanna grab some icecream?'
body = ''' yo
'''

em.add_alternative("""\
<!DOCTYPE html>
<html>
    <body>
        <h1 style="color:SlateGray;">This is an HTML Email!</h1>
    </body>
</html>
""", subtype='html')

em['From'] = email_sender
em['Subject'] = subj

context = ssl.create_default_context()

for i in range(2,5):

    email_receiver = sh.cell(row=i,column=2).value
    email_receiver_name = sh.cell(row=i,column=1).value
    email_receiver_name_prev = sh.cell(row=i-1, column=1).value

    # img_path = 'imagess/' + sh.cell(i,1).value + '.jpg'
    img_path = f'imagess/{email_receiver_name}.jpg'
    img = Image.open(img_path)
    img_fin = img.convert('RGB')
    # pdf_path = 'pic_to_pdf/' + sh.cell(i,1).value + '.pdf'
    pdf_path = f'pic_to_pdf/{email_receiver_name}.pdf'
    img_fin.save(pdf_path)
    print(pdf_path)
    # os.remove('pic_to_pdf/'+ email_receiver_name_prev + '.pdf')

    with open(pdf_path, 'rb') as f:
        file_data = f.read()
        file_name = f.name
    em.add_attachment(file_data, maintype='application',subtype='octet-stream',filename=file_name)

    em['To'] = email_receiver
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
    # with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
    #     smtp.starttls()
        smtp.login(email_sender, email_password)
        smtp.sendmail(email_sender, email_receiver, em.as_string())
        print(f'Message sent to {email_receiver_name}')
        os.remove(pdf_path)

    del em['To']

